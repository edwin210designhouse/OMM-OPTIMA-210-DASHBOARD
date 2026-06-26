"""
OMM Optima Container Dashboard — Auto-update script
Double-click this file to update the dashboard from Excel and push to GitHub.

Steps it runs automatically:
  1. Reads container schedule + unit matrix from Excel
  2. Injects fresh data into OMM_Optima_Dashboard.html
  3. Pushes to GitHub → live site updates in ~60 seconds
"""

import os
import re
import json
import subprocess
from datetime import datetime
import openpyxl

# ── CONFIG ──────────────────────────────────────────────────────────────────
SITE_DIR      = r"C:\Users\Edwin\OneDrive - Studio Snaidero Chicago\OMM_Optima Website"
EXCEL_FILE    = os.path.join(SITE_DIR, "Output_Optima_Container Matrix & Schedule .xlsx")
ONEDRIVE_HTML = os.path.join(SITE_DIR, "OMM_Optima_Dashboard.html")

# Work files live in C:\Temp — completely outside OneDrive, immune to sync corruption
TEMP_DIR      = r"C:\Temp\omm-optima"
HTML_FILE     = os.path.join(TEMP_DIR, "OMM_Optima_Dashboard.html")
TEMPLATE_FILE = os.path.join(TEMP_DIR, "OMM_Optima_Template.html")

AUTO_GIT   = True   # Set to False to skip the git push

SH_CTN  = "OMM_Overall Container Matrix"
SH_UNIT = "OMM_Unit Matrix"
# ────────────────────────────────────────────────────────────────────────────

def ensure_template():
    """
    Ensure a clean HTML template exists in C:\\Temp.
    On first run, copies from OneDrive. If OneDrive copy is corrupted, errors out clearly.
    """
    os.makedirs(TEMP_DIR, exist_ok=True)
    if os.path.exists(TEMPLATE_FILE):
        return True
    # Try to copy from OneDrive
    if os.path.exists(ONEDRIVE_HTML):
        with open(ONEDRIVE_HTML, 'rb') as f:
            raw = f.read().replace(b'\x00', b'')
        if b'DATA_START' in raw and b'</html>' in raw:
            with open(TEMPLATE_FILE, 'wb') as f:
                f.write(raw)
            print("  Template saved to C:\\Temp.")
            return True
    print("\nERROR: Could not find a valid HTML template.")
    print("Please ensure OMM_Optima_Dashboard.html is in the OMM_Optima Website folder.")
    return False

def fmt_date(val):
    """Format a datetime or string value for display (e.g. 'Aug 5, 2026')."""
    if val is None:
        return "—"
    if isinstance(val, datetime):
        months = "Jan Feb Mar Apr May Jun Jul Aug Sep Oct Nov Dec".split()
        return f"{months[val.month - 1]} {val.day}, {val.year}"
    s = str(val).strip()
    return s if s else "—"

def read_containers(ws):
    """Read container schedule rows. Data starts row 11, CTN# is col index 4."""
    containers = []
    for row in ws.iter_rows(min_row=11, values_only=True):
        if len(row) < 32:
            continue
        num = row[4]          # CONT. #  (integer 1-21)
        if not isinstance(num, int):
            continue
        containers.append({
            "num":          num,
            "status":       str(row[3]).strip() if row[3] else "PROJ",
            "week":         str(row[6]).strip()  if row[6] else "--",
            "floors":       str(row[7]).strip()  if row[7] else "--",
            "contCode":     str(row[10]).strip() if row[10] else "--",
            "contSeal":     str(row[11]).strip() if row[11] else "--",
            "loadDate":     fmt_date(row[13]),
            "shipDate":     fmt_date(row[16]),
            "vessel1":      str(row[17]).strip() if row[17] else "TBD",
            "vessel2":      str(row[19]).strip() if row[19] else "TBD",
            "etaLAX":       fmt_date(row[21]),
            "etaPhoenix":   fmt_date(row[27]),
            "deliveryDate": fmt_date(row[31]),
            "installDate":  fmt_date(row[39]) if len(row) > 39 else "--",
        })
    return sorted(containers, key=lambda x: x["num"])

# Room component columns: (display name, STS col index, CTN col index)
# Matches the OMM_Unit Matrix sheet layout (0-indexed from row tuple)
COMP_COLS = [
    ("Foyer",     7,  8),   ("Den",       9,  10),  ("Hall",     11, 12),
    ("Utility",   13, 14),  ("Closet",    15, 16),  ("Bev Ctr",  17, 18),
    ("Laundry",   19, 20),  ("Kitchen",   21, 22),  ("Dining",   23, 24),
    ("Great Rm",  25, 26),  ("Prim Bath", 27, 28),  ("Bath 2",   29, 30),
    ("Bath 3",    31, 32),  ("Bath 4",    33, 34),  ("Prim Bed", 35, 36),
    ("Bed 2",     37, 38),  ("Bed 3",     39, 40),  ("Bed 4",    41, 42),
]

def read_units(ws):
    """Read unit matrix rows. Data starts row 11."""
    units = []
    for row in ws.iter_rows(min_row=11, values_only=True):
        unit_num = row[3]     # UNIT column (e.g. 2006)
        if not isinstance(unit_num, int):
            continue
        status   = row[6]
        in_scope = (status != "NOT IN SCOPE")
        comps = []
        if in_scope:
            for name, si, ci in COMP_COLS:
                if si >= len(row) or ci >= len(row):
                    continue
                if row[si] == "YES" and isinstance(row[ci], (int, float)):
                    comps.append({"name": name, "ctn": int(row[ci])})
        ctns = sorted(set(c["ctn"] for c in comps))
        units.append({
            "floor":      row[1],
            "unit":       unit_num,
            "inScope":    in_scope,
            "combo":      row[4] if row[4] and row[4] != "NO" else None,
            "type":       row[5] or None,
            "status":     status,
            "components": comps,
            "ctns":       ctns,
        })
    return units

GITHUB_REMOTE = "https://github.com/edwin210designhouse/OMM-OPTIMA-210-DASHBOARD.git"

# Store .git OUTSIDE OneDrive so it can never be corrupted by sync
GIT_DIR = r"C:\Temp\omm-optima-git"

def git_env():
    """Git env vars: .git and work tree both in C:\\Temp, never touched by OneDrive."""
    e = os.environ.copy()
    e["GIT_DIR"] = GIT_DIR
    e["GIT_WORK_TREE"] = TEMP_DIR  # push from C:\Temp, not OneDrive
    return e

def git_setup():
    """Initialize git repo in C:\\Temp if not already done."""
    os.makedirs(GIT_DIR, exist_ok=True)
    env = git_env()
    def r(cmd):
        return subprocess.run(cmd, env=env, cwd=SITE_DIR, capture_output=True, text=True)
    r(["git", "init"])
    r(["git", "config", "user.email", "flores.edwin9271@gmail.com"])
    r(["git", "config", "user.name", "Edwin"])
    # Set remote (ignore error if already exists)
    subprocess.run(["git", "remote", "remove", "origin"], env=env, cwd=SITE_DIR, capture_output=True)
    r(["git", "remote", "add", "origin", GITHUB_REMOTE])

def git_push(commit_msg):
    """Push to GitHub using git stored in C:\\Temp (immune to OneDrive corruption)."""
    if not os.path.exists(GIT_DIR):
        git_setup()

    env = git_env()
    def r(cmd):
        return subprocess.run(cmd, env=env, cwd=SITE_DIR, capture_output=True, text=True)

    r(["git", "add", "-A"])
    rc = r(["git", "commit", "-m", commit_msg])
    if "nothing to commit" in (rc.stdout + rc.stderr):
        # Force a commit by touching the timestamp
        r(["git", "commit", "--allow-empty", "-m", commit_msg])

    rp = r(["git", "push", "-u", "origin", "main", "--force"])
    if rp.returncode == 0:
        return True, "pushed"

    # Push failed — re-init and retry once
    print("  Push failed, re-initializing git...")
    import shutil
    try:
        shutil.rmtree(GIT_DIR)
    except Exception:
        pass
    git_setup()
    r(["git", "add", "-A"])
    r(["git", "commit", "--allow-empty", "-m", commit_msg])
    rp2 = r(["git", "push", "-u", "origin", "main", "--force"])
    if rp2.returncode == 0:
        return True, "pushed (after reinit)"
    return False, rp2.stderr.strip()

def main():
    print("=" * 54)
    print("  OMM Optima Dashboard -- Auto-update")
    print("=" * 54)

    # ── Ensure template exists in C:\Temp ──────────────────────
    if not ensure_template():
        input("\nPress Enter to close...")
        return

    # ── Read Excel ──────────────────────────────────────────────
    print(f"\nReading: {os.path.basename(EXCEL_FILE)}")
    if not os.path.exists(EXCEL_FILE):
        print(f"\nERROR: Excel file not found at:\n  {EXCEL_FILE}")
        print("\nMake sure the Excel file is in the same folder as this script.")
        input("\nPress Enter to close...")
        return

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    except Exception as e:
        print(f"\nERROR opening Excel: {e}")
        input("\nPress Enter to close...")
        return

    if SH_CTN not in wb.sheetnames:
        print(f"\nERROR: Sheet '{SH_CTN}' not found.")
        print(f"Available sheets: {wb.sheetnames}")
        input("\nPress Enter to close...")
        return

    containers = read_containers(wb[SH_CTN])
    units      = read_units(wb[SH_UNIT])
    in_scope   = sum(1 for u in units if u["inScope"])
    print(f"  {len(containers)} containers | {in_scope} units in scope")

    if not containers:
        print("\nERROR: No containers found. Check that sheet name is correct.")
        input("\nPress Enter to close...")
        return

    # ── Build timestamp ─────────────────────────────────────────
    now = datetime.now()
    months = "Jan Feb Mar Apr May Jun Jul Aug Sep Oct Nov Dec".split()
    updated = f"{months[now.month - 1]} {now.day}, {now.year} at {now.strftime('%I:%M %p')}"

    # ── Build data block ────────────────────────────────────────
    data_block = (
        "// DATA_START\n"
        f"const CONTAINERS = {json.dumps(containers, ensure_ascii=False)};\n"
        f"const UNITS = {json.dumps(units, ensure_ascii=False)};\n"
        f'const UPDATED = "{updated}";\n'
        "// DATA_END"
    )

    # ── Update HTML (from C:\Temp template — immune to OneDrive) ──
    print(f"\nUpdating dashboard HTML...")
    if not os.path.exists(TEMPLATE_FILE):
        print(f"\nERROR: Template not found at:\n  {TEMPLATE_FILE}")
        input("\nPress Enter to close...")
        return

    try:
        with open(TEMPLATE_FILE, "rb") as f:
            raw = f.read().replace(b'\x00', b'')
        html = raw.decode("utf-8")
    except Exception as e:
        print(f"\nERROR reading template: {e}")
        input("\nPress Enter to close...")
        return

    new_html, count = re.subn(
        r"// DATA_START.*?// DATA_END",
        data_block,
        html,
        flags=re.DOTALL
    )

    if count == 0:
        print("\nERROR: Data markers not found in template.")
        input("\nPress Enter to close...")
        return

    try:
        # Write to C:\Temp (OneDrive never touches this)
        with open(HTML_FILE, "w", encoding="utf-8") as f:
            f.write(new_html)
        # Also copy to OneDrive folder so user can see it locally
        with open(ONEDRIVE_HTML, "w", encoding="utf-8") as f:
            f.write(new_html)
        print(f"  HTML updated successfully.")
    except Exception as e:
        print(f"\nERROR writing HTML: {e}")
        input("\nPress Enter to close...")
        return

    # ── Git push ─────────────────────────────────────────────────
    if AUTO_GIT:
        print("\nPushing to GitHub...")
        commit_msg = f"Auto-update {now.strftime('%Y-%m-%d %H:%M')} -- {len(containers)} containers"
        ok, msg = git_push(commit_msg)
        if ok:
            print(f"  OK   ({msg})")
            print(f"\nDone! Site will update within ~60 seconds.")
        else:
            print(f"\nERROR pushing to GitHub: {msg}")
            print("Close Excel and any other apps, then try again.")
    else:
        print(f"\nDone! (Git push skipped — AUTO_GIT = False)")

    print(f"\n  https://edwin210designhouse.github.io/OMM-OPTIMA-210-DASHBOARD/")
    input("\nPress Enter to close...")

if __name__ == "__main__":
    main()
